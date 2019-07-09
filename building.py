import collections
import csv
import openpyxl as pyxl
import numpy as np
from pathlib import Path

class Building:
    def __init__(self, name, lattitude, longitude, alias):
        self.name = name
        self.coords = (lattitude,longitude)
        self.alias = alias.split()
        self.distances = dict()

    def add_room(self, room, capacity):
        self.room = room
        self.capacity = capacity

    def distance_to(self, building):
        return self.distances[building.name]

    def __str__(self):
        return self.name

class Room:
    def __init__(self, name, building,capacity):
        self.name = name
        self.building = building
        self.capacity = capacity
        self.equipment = collections.defaultdict(int)

    def load_equipment(self, equip_list):
        equipments = equip_list.split(';')
        for equipment in equipments:
            count,name = equipment.replace(' ','').split('x',2)
            self.equipment[name] = count

class Rooms(collections.MutableMapping):
    def __init__(self, *args, **kwargs):
        self.store = dict()
        self.update(dict(*args, **kwargs))  # use the free update to set keys

    def __getitem__(self, key):
        return self.store[self._keytransform(key)]

    def __setitem__(self, key, value):
        self.store[self._keytransform(key)] = value

    def __delitem__(self, key):
        del self.store[self._keytransform(key)]

    def __iter__(self):
        return iter(self.store)

    def __len__(self):
        return len(self.store)

    def _keytransform(self, key):
        key = key.lower()
        key = key.replace('harry','h')
        key = key.replace('nursten','n')
        key = key.replace('foxh','fh')
        key = key.replace('jjthom','jjt')
        key = key.replace('hopk','hp')
        key = key.replace('wkh','wh')
        key = key.replace('pc','')
        key = key.replace('em','')
        key = key.replace('-','')
        key = key.replace(' ','')
        while 'l0' in key:
            key = key.replace('l0','l')
        #print(key)
        return key

def distance_matrix(buildings):
    """No longer used"""
    distances = []
    for x in buildings.values():
        distances.append([round(x.distance_to(y)) for y in buildings.values()])
    distances = np.array(distances)
    #print(distances)
    return distances

def get_building(room, buildings):
    for key,value in buildings.items():
        if room.startswith(key):
            return value

def load_buildings():
    p = Path('Data/buildings.csv')
    buildings = dict()
    with open(p,'r') as f:
        reader = csv.reader(f)
        next(reader,None)#skip header
        for name,lattitude,longitude,alias in reader:
            building = Building(name,lattitude,longitude,alias)
            buildings[name] = building
            for alias in building.alias:
                buildings[alias] = building
    load_distances(buildings)
    return buildings

def load_distances(buildings):
    p = Path('Data/buildings.csv')
    with open(p,'r') as f:
        reader = csv.reader(f)
        names = next(reader)[1:]
        for row in reader:
            for name,dist in zip(names,row[1:]):
                if dist == 'X':
                    buildings[row[0]].distances[name] = 0
                    break
                else:
                    buildings[row[0]].distances[name] = dist


def load_rooms(rooms_path, equip_path, buildings):
    rooms = Rooms()
    worksheet = pyxl.load_workbook(rooms_path,read_only = True, data_only = True).active
    for room_name,building_name,capacity in worksheet.iter_rows(min_row=2,values_only=True):
        building = buildings[building_name]
        rooms[room_name] = Room(room_name,building,capacity)
    load_equipment(equip_path, rooms)
    return rooms

def load_equipment(path, rooms):
    worksheet = pyxl.load_workbook(path,read_only = True, data_only = True).active
    for room_name,equipment_list in worksheet.iter_rows(min_row=2,max_col=2,values_only=True):
        try:
            rooms[room_name].load_equipment(equipment_list)
        except KeyError:
            print("No capacity listed for:",room_name)
