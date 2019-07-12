import os
import numpy as np
import openpyxl as pyxl
import sys
from building import *
from timeutils import *
from event import Event
from yattag import Doc,indent

class TimeSlot:
    def __init__(self,event):
        self.events = [event]

    def add(self,event):
        self.events.append(event)

    def sort(self):
        if self.events[0]:
            self.events.sort(key = lambda x: 1000 if x.date else min(x.weeks))

    def matches(self, event):
        if self.events[0]:
            return self.events[0].same_except_room(event)

class Timetable:
    def __init__(self, path):
        self.timetable = pyxl.load_workbook(path, read_only = True, data_only = True).active
        self.days = {day:[] for day in ("Mon","Tue","Wed","Thu","Fri","Sat","Sun")}
        self.events = []
        self.headers = []

    def generate_event_list(self,codes):
        eventlist = set()
        for row in self.timetable.iter_rows(min_row=2,values_only=True):
            module = str(row[1])
            if module and any(Timetable.matches(x,codes) for x in module.split(',')):
                eventlist.add(row[0])
        return list(eventlist)

    def load_events(self,eventlist):
        headers = set()
        #load events from excel
        for row in self.timetable.iter_rows(min_row=2,values_only=True):
            if row[0] in eventlist:
                if not (row[4] and row[5] and row[8]):
                    print("Warning: Missing date/time/room info for event id",row[0])
                    continue
                event = Event(row)
                for e in self.days[row[4]]:
                    if event.can_merge_with(e):
                        e.weeks.update(event.weeks)
                        event = None
                        break
                if event:
                    self.events.append(event)
                    self.days[event.day].append(event)
                    headers.update((event.start_time,event.end_time))
        for row in self.days.values():
            row.sort()
        self.headers = sorted(headers)
        for event in self.events:
            event.col_span(self.headers)

    def create_timetable(self):
        self.rows = []
        for day,row in self.days.items():
            row.sort()
            cell_list = [[None] * (len(self.headers) - 1)]
            for event in row:
                slotted = False
                i = 0
                while not slotted:
                    cells = cell_list[i]
                    #slot into time table if there is space
                    if cells[event.start] == None:
                        cells[event.start]= TimeSlot(event)
                        for i in range(event.start+1,event.end):
                            cells[i] = TimeSlot(None)
                        slotted = True
                    #if different type/group add to time slot
                    elif cells[event.start].matches(event):
                        cells[event.start].add(event)
                        slotted = True
                    else:#if in conflict add a new row
                        i+=1
                        if len(cell_list) == i:
                            cell_list.append([None] * (len(self.headers) -1))

            for cells in cell_list:
                [c.sort() for c in cells if c]
                self.rows.append((day,cells))

    def list_modules(self):
        modules = set()
        for row in self.timetable.iter_rows(min_row=2,values_only=True):
            if row[1]:
                modules.update(str(row[1]).split(","))
        modules = sorted(modules)
        return modules

    def render_html(self, headerbg = '#C0C0C0', headerfont = '#000000', cellbg = '#00FFFF', cellfont = '#000000'):
        self.create_timetable()
        doc, tag, text, line = Doc().ttl()
        doc.asis('<!DOCTYPE html>')
        with tag('html'):
            with tag('head'):
                line('title','Time Table Test')
            with tag('body'):
                with tag('table', cellspacing=0, border=1):
                    doc.asis('<!-- START COLUMNS HEADERS-->')
                    with tag('tr'):
                        line('td','')
                        for header in self.headers[:-1]:#last header marks end
                            with tag('td', bgcolor=headerbg, colspan=1):
                                line('font', str(minute_to_time(header))[:5], color=headerfont)
                    doc.asis('<!-- END COLUMNS HEADERS-->')

                    for day,row in self.rows:
                        doc.asis('<!--START ROW '+day+'-->')
                        with tag('tr'):
                            with tag('td', bgcolor=headerbg, rowspan=1):
                                line('font',day,color=headerfont)
                            for cell in row:
                                if cell:
                                    event = cell.events[0]
                                    if not event:
                                        continue
                                    with tag('td', bgcolor=cellbg, colspan=event.span, rowspan=1):
                                        text = event.modules_str() + " " + event.type
                                        text += " " + event.groups_str()
                                        Timetable.create_table(text,tag,line,cellbg)
                                        for e in cell.events:
                                            Timetable.create_table(e.room,tag,line,cellbg)
                                            Timetable.create_table(e.weeks_str(),tag,line,cellbg)
                                else:#empty cell
                                    with tag('td'):
                                        doc.asis('&nbsp')
                        doc.asis('<!--END ROW '+day+'-->')


        return indent(doc.getvalue())

    def create_table(text,tag,line,cellbg):
        with tag('table', bgcolor=cellbg, cellspacing=0, border=0, width='100%'):
            with tag('tr'):
                line('td', text, align='left')

    def matches(string, patterns):
        return any(string.startswith(pattern) for pattern in patterns)

if __name__ == "__main__":
    files = ["Room List.xlsx", "Roomequip.xlsx", "Timetable2018-19.xlsx","buildings.csv"]
    paths = [os.path.join("Data",x) for x in files]

    #buildings = load_buildings(paths[3])
    #rooms = load_rooms(paths[0], paths[1], buildings)

    #timetable = pyxl.load_workbook(paths[2]).active
    t = Timetable(paths[2])
    eventlist = t.generate_event_list([sys.argv[1]])
    t.load_events(eventlist)

    out = t.render_html()
    with open ('timetable.html','w') as f:
        f.write(out)
    #get_info(paths[2])
